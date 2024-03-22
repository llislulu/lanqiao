from openpyxl import load_workbook
from collections import Counter

# 从Excel文件中读取数据并解析
workbook = load_workbook('test.xlsx')
sheet = workbook.active
names = []

# 从第二行开始读取数据
for row in sheet.iter_rows(min_row=4, values_only=True):
    names.extend(row[1:])  # 提取所有名字

# 使用Counter对象统计每个名字出现的次数
name_counts = Counter(names)

# 输出结果
for name in name_counts.keys():
    print(f"{name}: {name_counts[name]}次")


from openpyxl import Workbook

# 创建一个新的Excel工作簿
workbook = Workbook()

# 获取工作簿的活动表
sheet = workbook.active

# 将结果写入表格
for idx, (name, count) in enumerate(name_counts.items(), start=1):
    sheet[f'A{idx}'] = name  # 将名字写入第一列
    sheet[f'B{idx}'] = count  # 将出现次数写入第二列

# 保存工作簿到文件
workbook.save('周六晚.xlsx')